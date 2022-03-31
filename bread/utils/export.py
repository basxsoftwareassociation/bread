import html
import io
import re

import htmlgenerator as hg
from django.conf import settings
from django.http import HttpResponse
from django.utils.html import strip_tags


def html_to_pdf(html, as_http_response=False, name=None):
    """
    Renders html to PDF

        html: html string
        as_http_response: If True return PDF as HttpResponse with attachment
        name: Filename without extension, only used when as_http_response == True
        returns: PDF content or HttpResponse with attachment
    """
    # weasyprint is an extra dependency
    import weasyprint

    def url_fetcher(url, timeout=10, ssl_context=None):
        return weasyprint.default_url_fetcher(
            "file://" + url.replace("file://", settings.BASE_DIR), timeout, ssl_context
        )

    ret = weasyprint.HTML(string=html, base_url="", url_fetcher=url_fetcher).write_pdf()
    if as_http_response:
        ret = HttpResponse(ret, content_type="application/pdf")
        ret["Content-Disposition"] = f"inline; filename={name}.pdf"
    return ret


def generate_excel(rows, columns):
    """
    columns: dict with {<columnname>: formatting_function(row)}
    """
    import openpyxl
    from openpyxl.styles import Font

    workbook = openpyxl.Workbook()
    workbookcolumns = workbook.active.iter_cols(
        min_row=1, max_col=len(columns) + 1, max_row=len(rows) + 1
    )
    newline_regex = re.compile(
        r"<\s*br\s*/?\s*>"
    )  # replace HTML line breaks with newlines
    for columnname, columndata in zip(columns, workbookcolumns):
        columndata[0].value = html.unescape(
            strip_tags(newline_regex.sub(r"\n", str(columnname)))
        )
        columndata[0].font = Font(bold=True)
        for i, cell in enumerate(columndata[1:]):
            value = columns[columnname](rows[i])
            if isinstance(value, hg.BaseElement):
                value = hg.render(value, {})
            cleaned = html.unescape(
                strip_tags(newline_regex.sub(r"\n", str(value or "")))
            )
            cell.value = cleaned
    return workbook


def xlsxresponse(workbook, title, filters=True):
    """
    Returns workbook as a downloadable file

        workbook: openpyxl workbook
        title: filename without extension
        returns: HttpResponse with attachment
    """
    workbook = prepare_excel(workbook, filters)
    buf = io.BytesIO()
    workbook.save(buf)
    buf.seek(0)
    response = HttpResponse(buf.read(), content_type="application/vnd.ms-excel")
    response["Content-Disposition"] = f'attachment; filename="{title}.xlsx"'
    return response


def prepare_excel(workbook, filters=True):
    """
    Formats the excel a bit in order to be displayed nicely

        workbook: openpyxl workbook
        filters: If True enable excel filtering headers
        returns: formated workbook
    """
    # openpyxl is an extra requirement
    from openpyxl.styles import Alignment

    for worksheet in workbook:
        # estimate column width
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            adjusted_width = (max_length + 3) * 1.2
            worksheet.column_dimensions[column].width = min([adjusted_width, 50])

            # enable excel filters
            if filters is True:
                worksheet.auto_filter.ref = f"A1:{column}1"

        # enable word wrap
        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)
                if isinstance(cell.value, str):
                    cell.value = cell.value.strip()
                    if cell.value.isdigit():
                        cell.value = int(cell.value)

    return workbook
