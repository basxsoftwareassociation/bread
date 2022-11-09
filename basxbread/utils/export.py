import html
import io
import re

import htmlgenerator as hg
from django.http import HttpResponse
from django.utils.html import strip_tags


def generate_excel(rows, columns):
    """
    columns: dict with {<columnname>: formatting_function(row)}
    """
    import openpyxl
    from openpyxl.styles import Font

    workbook = openpyxl.Workbook()
    workbookcolumns = workbook.active.iter_cols(
        min_row=1, max_col=len(columns) + 1, max_row=len(rows) + 1
    )
    newline_regex = re.compile(
        r"<\s*br\s*/?\s*>"
    )  # replace HTML line breaks with newlines
    for columnname, columndata in zip(columns, workbookcolumns):
        columndata[0].value = html.unescape(
            strip_tags(newline_regex.sub(r"\n", str(columnname)))
        )
        columndata[0].font = Font(bold=True)
        for i, cell in enumerate(columndata[1:]):
            value = columns[columnname](rows[i])
            if isinstance(value, hg.BaseElement):
                value = hg.render(value, {})

            cleaned = html.unescape(
                strip_tags(newline_regex.sub(r"\n", str(value or "")))
                .replace("\n\n", "\n")
                .strip()
            )
            cell.value = cleaned
    return workbook


def xlsxresponse(workbook, title, filters=True):
    """
    Returns workbook as a downloadable file

        workbook: openpyxl workbook
        title: filename without extension
        returns: HttpResponse with attachment
    """
    workbook = prepare_excel(workbook, filters)
    buf = io.BytesIO()
    workbook.save(buf)
    buf.seek(0)
    response = HttpResponse(buf.read(), content_type="application/vnd.ms-excel")
    response["Content-Disposition"] = f'attachment; filename="{title}.xlsx"'
    return response


def prepare_excel(workbook, filters=True):
    """
    Formats the excel a bit in order to be displayed nicely

        workbook: openpyxl workbook
        filters: If True enable excel filtering headers
        returns: formated workbook
    """
    # openpyxl is an extra requirement
    from openpyxl.styles import Alignment

    for worksheet in workbook:
        # estimate column width
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            adjusted_width = (max_length + 3) * 1.2
            worksheet.column_dimensions[column].width = min([adjusted_width, 50])

            # enable excel filters
            if filters is True:
                worksheet.auto_filter.ref = f"A1:{column}1"

        # enable word wrap
        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)
                if isinstance(cell.value, str):
                    cell.value = cell.value.strip()
                    if cell.value.isdigit():
                        cell.value = int(cell.value)

    return workbook
